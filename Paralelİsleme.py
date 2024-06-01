from textblob import TextBlob
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import matplotlib.pyplot as plt

# Excel dosyasını yükle
dosya_yolu = "/Users/ibrahimcomert/Desktop/YoutubeYorum.xlsx"  # Excel dosyasının yolu
workbook = load_workbook(dosya_yolu)
sheet = workbook.active

# Yorumları çek
yorumlar = []
for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):  # İlgili sütundan yorumları çek (ör. A sütunu)
    if row[1]:  # Yorumun None olmadığından emin ol
        yorumlar.append(row[1])

# Çeviri ve duygu analizi fonksiyonu
def ceviri_ve_analiz(yorum):
    try:
        # Yorumu İngilizce'ye çevir
        ingilizce_yorum = GoogleTranslator(source='auto', target='en').translate(yorum)
        # Duygu analizi yap
        analiz = TextBlob(str(ingilizce_yorum))
        return analiz.sentiment.polarity, True
    except Exception as e:
        print(f"Hata oluştu: {e}")
        return None, False

# Yorumları paralel olarak işleme
duygu_skorları = []
toplam_kayit_sayisi = len(yorumlar)
basarili_kayit_sayisi = 0
hatali_kayit_sayisi = 0

with ThreadPoolExecutor(max_workers=10) as executor:  # 10 iş parçacığı kullanarak
    future_to_yorum = {executor.submit(ceviri_ve_analiz, yorum): yorum for yorum in yorumlar}
    for future in as_completed(future_to_yorum):
        result, success = future.result()
        if success:
            duygu_skorları.append(result)
            basarili_kayit_sayisi += 1
        else:
            hatali_kayit_sayisi += 1

# Çeviri istatistiklerini ekrana yazdırma
print(f"Toplam kayıt sayısı: {toplam_kayit_sayisi}")
print(f"Başarılı çevrilen kayıt sayısı: {basarili_kayit_sayisi}")
print(f"Hatalı çevrilen kayıt sayısı: {hatali_kayit_sayisi}")

# Duygu skorlarını kategorilere ayırma
pozitif = sum(1 for skor in duygu_skorları if skor > 0)
negatif = sum(1 for skor in duygu_skorları if skor < 0)
notr = sum(1 for skor in duygu_skorları if skor == 0)

# Grafik oluşturma
labels = ['Pozitif', 'Negatif', 'Nötr']
sizes = [pozitif, negatif, notr]
colors = ['green', 'red', 'gray']
explode = (0.1, 0, 0)  # Sadece pozitif duyguları vurgulamak için

# Eğer tüm duygu skorları 0 ise (yorumlar düzgün yüklenmediyse), sizes içinde NaN değer olabilir. Bunu kontrol edelim.
if sum(sizes) == 0:
    print("Yorumlar veya duygu skorları doğru yüklenemedi.")
else:
    plt.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%', startangle=140)
    plt.axis('equal')  # Daireyi çember yapmak için
    plt.title('Yorum Duyguları Analizi')
    plt.show()
